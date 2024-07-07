# -*- coding: utf-8 -*-
"""
Updated 1/14/2021
Updated from dashboard_code by Aurora Namnum
Contains functions necessary to retreive data with the ENERGY STAR API. Dashboard creation is done separately.
"""

import xml.etree.ElementTree as ET
import requests
from requests.auth import HTTPBasicAuth
import pandas as pd
from datetime import datetime
from datetime import timedelta
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import win32com.client as win32
import shutil
import os
import ctypes

pd.options.mode.chained_assignment = None  # default='warn'


#%%% PRESETS

def get_current_date():
    """Gets current date

    :return: strings of current year, month, and day
    """
    # PRESET current date
    todaysDate = {"day":datetime.now().day, 
                  "month":datetime.now().month,
                  "year":datetime.now().year}
    year_curr = str(todaysDate['year'])
    month_curr = str(todaysDate['month'])
    if len(month_curr) == 1:
        month_curr = '0' + month_curr
    day_curr = str(todaysDate['day'])
    if len(day_curr) == 1:
        
        day_curr = '0' + day_curr
    return year_curr,month_curr,day_curr

def set_preconditions():
    """
    """
    # PRESET metrics list
    metrics_list = ['score','siteIntensity','totalLocationBasedGHGEmissions']
    
    # PRESETS for API
    api_url_base = 'https://portfoliomanager.energystar.gov/ws/'
    username = 'Ithaca2030_dashboard'
    pswd = '1/2by2050'
    
    # PRESET lists to build dataframes
    prop1_temp=[]
    acc_temp=[]
    name_temp=[]
    propType_temp=[]
    yearBuilt_temp=[]
    add1_temp=[]
    city_temp=[]
    country_temp=[]
    postCode_temp=[]
    state_temp=[]
    prop2_temp=[]
    gfa_temp=[]
    units1_temp=[]
    prop3_temp=[]
    met_year_temp=[]
    score_temp=[]
    eui_temp=[]
    ghg_temp=[]
    prop4_temp=[]
    mtr1_temp=[]
    mtrType_temp=[]
    units2_temp=[]
    prop5_temp=[]
    mtr2_temp=[]
    month_temp=[]
    cons_temp=[]
    cost_temp=[]
    
    # PRESET maximum entries in a meter
    maxEntries = 0
    
    return metrics_list,api_url_base,username,pswd,prop1_temp,acc_temp,name_temp,propType_temp,yearBuilt_temp,add1_temp,city_temp,country_temp,postCode_temp,state_temp,prop2_temp,gfa_temp,units1_temp,prop3_temp,met_year_temp,score_temp,eui_temp,ghg_temp,prop4_temp,mtr1_temp,mtrType_temp,units2_temp,prop5_temp,mtr2_temp,month_temp,cons_temp,cost_temp,maxEntries
    


#%% ACCOUNT ID

def get_connected_accounts(api_url_base,username,pswd):
    """Gets IDs for accounts that have been shared
    :param username: Username for admin account
    :type username: str
    :param pswd: password for admin account
    :type pswd: str

    :return: string containing account information in XML format
    """

    api_url = api_url_base + 'customer/list'
    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd))

    if response.status_code == 200:
        return response.content #xml file
    else:
        return None


def xmltodict_acc_cnxn(api_url_base,username,pswd):
    
    acc_connections_xml = get_connected_accounts(api_url_base,username,pswd)
    if acc_connections_xml is not None:
        acc_connections = ET.fromstring(acc_connections_xml)
        
        accounts = {}
        for link in acc_connections.iter('link'): 
            accounts.update({link.attrib['id']:""})
        
    return accounts

def get_account_info(api_url_base,username,pswd,acc_id):
    """Gets account information for admin account
    :param username: Username for admin account
    :type username: str
    :param pswd: password for admin account
    :type pswd: str

    :return: string containing account information in XML format
    """

    api_url = api_url_base + 'customer/' + acc_id
    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd))

    if response.status_code == 200:
        return response.content  # xml file
    else:
        return None

def xmltodict_acc_info(api_url_base,username,pswd,accounts):    
    for acc_id in accounts:
        acc_info_xml = get_account_info(api_url_base,username,pswd,acc_id)
        acc_info = ET.fromstring(acc_info_xml)

        accounts[acc_id] = (acc_info[0].text)
        
    return accounts



#%% ADMIN ACCOUNT INFO

global testvariablefrommike
testvariablefrommike = 0
print('defined variable')


def get_admin_account(api_url_base,username,pswd,prop_id):
    """
    """

    api_url = api_url_base + 'idHierarchy/property/' + str(prop_id)
    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd))
    global testvariablefrommike
    testvariablefrommike += 1  
    print(response.status_code, 'testvariablefrommike is ', testvariablefrommike)
    if response.status_code == 200:
        return response.content  # xml file
    else:
        return None

def xmltodict_admin_acc(api_url_base,username,pswd,inputs_df):    
    """

    Parameters
    ----------
    api_url_base : TYPE
        DESCRIPTION.
    username : TYPE
        DESCRIPTION.
    pswd : TYPE
        DESCRIPTION.
    accounts : TYPE
        DESCRIPTION.

    Returns
    -------
    TYPE
        DESCRIPTION.

    """
    
    data = {}
    for prop_id in inputs_df['Property ID (from ESPM Profile)']:
        admin_info_xml = get_admin_account(api_url_base,username,pswd,prop_id)
        admin_info = ET.fromstring(admin_info_xml)      
        try:
            data[admin_info[0].text].update({str(prop_id):{}})
        except:
            data.update({admin_info[0].text:{}})
            data[admin_info[0].text].update({str(prop_id):{}})
    
    return data


def get_district_name():
    """
    """
    inputs_df = pd.read_excel("inputs.xlsx", sheet_name="inputs", header=None, usecols="C")
    district_name = inputs_df[2][2]
    
    return district_name

def get_which_properties():
    """
    """
    inputs_df = pd.read_excel("inputs.xlsx", sheet_name="inputs", header=None, usecols="C")
    which_properties = inputs_df[2][3]
    
    return which_properties

def get_which_metrics():
    """
    """
    inputs_df = pd.read_excel("inputs.xlsx", sheet_name="inputs", header=None, usecols="C")
    which_metrics = inputs_df[2][4]
    
    return which_metrics

def get_district_eui_wui():
    """
    """
    inputs_df = pd.read_excel("inputs.xlsx", sheet_name="inputs", header=None, usecols="C")
    district_eui = inputs_df[2][5]
    district_wui = inputs_df[2][6]
    
    return district_eui, district_wui


#%% INPUTS
    
def get_inputs():
    """
    """
    try:
        inputs_df = pd.read_excel("inputs.xlsx", sheet_name="inputs", header=0, skiprows=[0,1,2,3,4,5,6,7], usecols=lambda x: 'Unnamed' not in x, dtype={'Property ID (from ESPM Profile)': str, 'Baseline EUI [kBtu/ft2]': int, 'Baseline WUI [gal/ft2]': int})
    except:
        ctypes.windll.user32.MessageBoxW(0, "Check the inputs sheet - did you leave any of the baseline values blank, or change the template in any way?","Error", 0)
        exit()
    return inputs_df

def get_transp_inputs(prop_id):
    """
    """
    try:
        temp = pd.read_excel("inputs.xlsx", sheet_name=prop_id, header=None, usecols="E")
        baselineT = temp[4][1]
        transpSplit = pd.read_excel("inputs.xlsx", sheet_name=prop_id, header=0, skiprows=[0,1,2,3], usecols=[1,2,3,4,5])
        transpSplit = transpSplit.dropna()
        transpTotal = pd.read_excel("inputs.xlsx", sheet_name=prop_id, header=0, skiprows=[0,1,2,3], usecols=[7,8])
        transpTotal = transpTotal.dropna()
        new_which_metrics = which_metrics
    except:
        new_which_metrics = "Energy & Water"
        baselineT = 0
        transpSplit = pd.DataFrame()
        transpTotal = pd.DataFrame()
    
    return baselineT, transpSplit, transpTotal, new_which_metrics
    
def get_transp_agg_inputs():    
    """
    """
    try:
        temp = pd.read_excel("inputs.xlsx", sheet_name="ithaca - agg transportation", header=None, usecols="E")
        baselineT = temp[4][1]
        transpSplit = pd.read_excel("inputs.xlsx", sheet_name="ithaca - agg transportation", header=0, skiprows=[0,1,2,3], usecols=[1,2,3,4,5])
        transpSplit = transpSplit.dropna()
        transpTotal = pd.read_excel("inputs.xlsx", sheet_name="ithaca - agg transportation", header=0, skiprows=[0,1,2,3], usecols=[7,8])
        transpTotal = transpTotal.dropna()
        new_which_metrics = which_metrics
    except:
        new_which_metrics = "Energy & Water"
        baselineT = pd.DataFrame()
        transpSplit = pd.DataFrame()
        transpTotal = pd.DataFrame()
        
    return baselineT, transpSplit, transpTotal, new_which_metrics


#%% ACCEPT REQUESTS

def get_account_requests(api_url_base,username,pswd):
        
    api_url = api_url_base + 'connect/account/pending/list'
    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd))
    
    if response.status_code == 200:
        return response.content #xml file
    elif response.status_code == 502:
        ctypes.windll.user32.MessageBoxW(0, "Portfolio Manager may be temporarily offline. Try again later.","Error", 0)
        exit()            
    else:
        return None

def xmltodict_acc_rqsts(api_url_base,username,pswd):    
    acc_requests_xml = get_account_requests(api_url_base,username,pswd)
    acc_requests = ET.fromstring(acc_requests_xml)

    
    acc_rqsts = []
    for acc in acc_requests.iter('account'): 
        acc_rqsts.append(acc[0].text)
        
    return acc_rqsts

def get_property_requests(api_url_base,username,pswd):
        
    api_url = api_url_base + 'share/property/pending/list'
    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd))
    
    if response.status_code == 200:
        return response.content #xml file
    else:
        return None

def xmltodict_prop_rqsts(api_url_base,username,pswd):    
    prop_requests_xml = get_property_requests(api_url_base,username,pswd)
    prop_requests = ET.fromstring(prop_requests_xml)
    
    prop_rqsts = []
    for prop in prop_requests.iter('property'): 
        prop_rqsts.append(prop[0].text)
        
    return prop_rqsts

def get_meter_requests(api_url_base,username,pswd):
        
    api_url = api_url_base + 'share/meter/pending/list'
    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd))
    
    if response.status_code == 200:
        return response.content #xml file
    else:
        return None

def xmltodict_mtr_rqsts(api_url_base,username,pswd):    
    mtr_requests_xml = get_meter_requests(api_url_base,username,pswd)
    mtr_requests = ET.fromstring(mtr_requests_xml)
    
    mtr_rqsts = []
    for mtr in mtr_requests.iter('meter'): 
        mtr_rqsts.append(mtr[0].text)
        
    return mtr_rqsts


def accept_acct_requests(api_url_base,acc_rqsts,username,pswd):
    # Create XML body
    sharingResponse = ET.Element('sharingResponse')
    action = ET.SubElement(sharingResponse, 'action')
    action.text = 'Accept'
    note = ET.SubElement(sharingResponse, 'note')
    note.text = 'Your connection request has been verified and accepted.'
    xml = ET.tostring(sharingResponse)
   
    headers={'Content-Type': 'application/xml'}
    for iD in acc_rqsts:
        api_url = api_url_base + 'connect/account/' + iD
        requests.post(api_url, data=xml, headers=headers, auth=HTTPBasicAuth(username,pswd))
    

def accept_prop_requests(api_url_base,prop_rqsts,username,pswd):
    # Create XML body
    sharingResponse = ET.Element('sharingResponse')
    action = ET.SubElement(sharingResponse, 'action')
    action.text = 'Accept'
    note = ET.SubElement(sharingResponse, 'note')
    note.text = 'Your connection request has been verified and accepted.'
    xml = ET.tostring(sharingResponse)
   
    headers={'Content-Type': 'application/xml'}
    
    for iD in prop_rqsts:
        api_url = api_url_base + 'share/property/' + iD
        requests.post(api_url, data=xml, headers=headers, auth=HTTPBasicAuth(username,pswd))
        

def accept_mtr_requests(api_url_base,mtr_rqsts,username,pswd):
    # Create XML body
    sharingResponse = ET.Element('sharingResponse')
    action = ET.SubElement(sharingResponse, 'action')
    action.text = 'Accept'
    note = ET.SubElement(sharingResponse, 'note')
    note.text = 'Your connection request has been verified and accepted.'
    xml = ET.tostring(sharingResponse)
   
    headers={'Content-Type': 'application/xml'}
       
    for iD in mtr_rqsts:
        api_url = api_url_base + 'share/meter/' + iD
        requests.post(api_url, data=xml, headers=headers, auth=HTTPBasicAuth(username,pswd))


#%% PROPERTY ID

def get_connected_properties(api_url_base,username,pswd,acc_id):
    """Gets IDs for properties that have been shared
    :param username: Username for admin account
    :type username: str
    :param pswd: password for admin account
    :type pswd: str
    :param accID: account ID for customer account
    :type accID: str
    
    :return: string containing property information in XML format
    """
    api_url = api_url_base + 'account/' + acc_id + '/property/list'
    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd))

    if response.status_code == 200:
        return response.content #xml file
    else:
        return None

def xmltodict_prop_cnxn(api_url_base,username,pswd,data,acc_id):    
    prop_connections_xml = get_connected_properties(api_url_base,username,pswd,acc_id)
    
    if prop_connections_xml is not None:
        prop_connections = ET.fromstring(prop_connections_xml)
        
        for link in prop_connections.iter('link'): 
            data[acc_id][link.attrib['id']] = {}
        
    return data


#%%% PROPERTY INFO (NAME, SF, YEAR BULIT)

def get_prop_info(api_url_base,username,pswd,prop_id):
    """
    """
    api_url = api_url_base + 'property/' + str(prop_id)
    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd))
    
    if response.status_code == 200:
        return response.content
    else:
        return None
    
def xmltodict_prop_info(api_url_base,username,pswd,data,prop_id):
    prop_info_xml = get_prop_info(api_url_base,username,pswd,prop_id)
    if prop_info_xml is not None:
        prop_info = ET.fromstring(prop_info_xml)
        
        data[acc_id][prop_id] = {"name":prop_info[0].text, 
            "propertyType":prop_info[4].text,
            "yearBuilt":prop_info[5].text,
            "meters":{},
            "metrics":{}}
        
        for info in prop_info.iter('address'):
            data[acc_id][prop_id].update({info.tag:info.attrib})
        
        for info in prop_info.iter('grossFloorArea'):
            data[acc_id][prop_id].update({info.tag:info.attrib})
            data[acc_id][prop_id]['grossFloorArea'].update({'value':prop_info[6][0].text})
        if data[acc_id][prop_id]['grossFloorArea']['units'] == "Square Meters":
            data[acc_id][prop_id]['grossFloorArea']['units'] = 'Square Feet'
            data[acc_id][prop_id]['grossFloorArea']['value'] = str(int(data[acc_id][prop_id]['grossFloorArea']['value'])*10.764)
        
    return data


#%%% METER ID
    
def get_connected_meters(api_url_base,username,pswd,prop_id):
    """
    """
    api_url = api_url_base + 'property/' + str(prop_id) + '/meter/list?myAccessOnly=true'
    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd))
    
    if response.status_code == 200:
        return response.content
    else:
        return None
    
def xmltodict_mtr_cnxn(api_url_base,username,pswd,data,acc_id,prop_id):
    mtr_connections_xml = get_connected_meters(api_url_base,username,pswd,prop_id)
    
    if mtr_connections_xml is not None:
        mtr_connections = ET.fromstring(mtr_connections_xml)
        
        for link in mtr_connections.iter('link'): 
            data[acc_id][prop_id]['meters'][link.attrib['link'].split('/meter/',1)[1]] = {}

    return data

#%% METER INFO (TYPE, EARLIEST DATE, UNITS)

def get_mtr_info(api_url_base,username,pswd,mtr_id):
    """
    """
    api_url = api_url_base + 'meter/' + mtr_id
    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd))
    
    if response.status_code == 200:
        return response.content
    else:
        return None
    
def xmltodict_mtr_info(api_url_base,username,pswd,data,acc_id,prop_id,mtr_id):
    mtr_info_xml = get_mtr_info(api_url_base,username,pswd,mtr_id)
    if mtr_info_xml is not None:
        mtr_info = ET.fromstring(mtr_info_xml)
        
        data[acc_id][prop_id]['meters'][mtr_id] = {"type":mtr_info[1].text, 
            "units":mtr_info[4].text,
            "firstBillDate":mtr_info[5].text,
            "consumptionData":{},
            "consumptionData_calendarized":{}}

    return data

#%%% METRICS

def get_metrics(api_url_base,username,pswd,prop_id,year,metric):
    """Gets consumption data for a given meter
            :param username: Username for admin account
            :type username: str
            :param pswd: password for admin account
            :type pswd: str
            :param propertyID: property ID for customer retrieving data from
            :type propertyID: str

            :return: string containing requested in XML format
            """
    
    headers_for_metrics = {'PM-Metrics': metric}
    
    api_url = api_url_base + 'property/' + str(prop_id) + '/metrics?year=' + year + '&month=12' + '&measurementSystem=EPA'

    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd), headers=headers_for_metrics)

    if response.status_code == 200:
        return response.content
    else:
        return None

def xmltodict_metrics(api_url_base,username,pswd,data,acc_id,prop_id,year,metric):    
    metrics_xml = get_metrics(api_url_base,username,pswd,prop_id,year,metric)
    if metrics_xml is not None:
        metrics = ET.fromstring(metrics_xml)
        if metrics[0][0].text is not None:
            if metric == 'totalLocationBasedGHGEmissions':
                data[acc_id][prop_id]["metrics"][year].update({metric:str(2204.62*float(metrics[0][0].text))})
            else:
                data[acc_id][prop_id]["metrics"][year].update({metric:str(float(metrics[0][0].text))})
        else:
            data[acc_id][prop_id]["metrics"][year].update({metric:'N/A'})
    
    return data

#%%% RESET EARLIEST DATE
def reset_date(earliestDate):
    if '01-01' in earliestDate:
        year = earliestDate[0:4]
    else:
        year = str(int(earliestDate[0:4]) - 1)
    
    return year

#%%% CONSUMPTION DATA
def get_consumption(api_url_base,username,pswd,mtr_id):
    """Gets consumption data for a given meter
    :param username: Username for admin account
    :type username: str
    :param pswd: password for admin account
    :type pswd: str
    :param meterID: meter ID for customer retrieving data from
    :type meterID: str

    :return: string containing consumption information in XML format
    """

    api_url = api_url_base + 'meter/' + mtr_id + '/consumptionData'

    response = requests.get(api_url, auth=HTTPBasicAuth(username,pswd))

    if response.status_code == 200:
        return response.content #xml file
    else:
        return None

def xmltodict_consumption(api_url_base,username,pswd,data,acc_id,prop_id,mtr_id):    
    consumption_xml = get_consumption(api_url_base,username,pswd,mtr_id)
    if consumption_xml is not None:
        consumption = ET.fromstring(consumption_xml)
        
        x=1
        for start in consumption.iter('startDate'): 
            if len(str(x)) == 1:
                y = '0' + str(x)
            else:
                y = str(x)
            data[acc_id][prop_id]["meters"][mtr_id]["consumptionData"].update({'entry'+y:{}})
            data[acc_id][prop_id]["meters"][mtr_id]["consumptionData"]['entry'+y].update({'startDate':start.text})
            x+=1
        x=1
        for end in consumption.iter('endDate'): 
            if len(str(x)) == 1:
                y = '0' + str(x)
            else:
                y = str(x)
            data[acc_id][prop_id]["meters"][mtr_id]["consumptionData"]['entry'+y].update({'endDate':end.text})
            x+=1
        x=1
        for cons in consumption.iter('usage'): 
            if len(str(x)) == 1:
                y = '0' + str(x)
            else:
                y = str(x)
            data[acc_id][prop_id]["meters"][mtr_id]["consumptionData"]['entry'+y].update({'value':float(cons.text)})
            x+=1
        x=1
        for cost in consumption.iter('cost'): 
            if len(str(x)) == 1:
                y = '0' + str(x)
            else:
                y = str(x)
            data[acc_id][prop_id]["meters"][mtr_id]["consumptionData"]['entry'+y].update({'cost':float(cost.text)})
            x+=1
            
    return data    

def calendarize_consumption(data,inputs_df,which_metrics):
    billFreq = {'Annually':12,'Semi-Annually':6,'Quarterly':3,'Monthly':1}
    for acc_id in data:
        for prop_id in data[acc_id]:
            if data[acc_id][prop_id]:
                if which_metrics != "Energy":
                    try:
                        wFreq = billFreq[inputs_df.loc[inputs_df.index[inputs_df['Property ID (from ESPM Profile)'] == prop_id][0],'Frequency of Water Bills']]
                    except:
                        ctypes.windll.user32.MessageBoxW(0, "Check the inputs sheet - your water bill frequency value was not accepted.","Error", 0)
                if data[acc_id][prop_id]['meters']:
                    for mtr_id in data[acc_id][prop_id]['meters']:
                        for entry in data[acc_id][prop_id]['meters'][mtr_id]['consumptionData']:
                            #get start and end dates for the entry    
                            start = data[acc_id][prop_id]['meters'][mtr_id]['consumptionData'][entry]['startDate']    
                            end = data[acc_id][prop_id]['meters'][mtr_id]['consumptionData'][entry]['endDate']
                            #convert start and end dates to datetime format
                            start = datetime.strptime(start, '%Y-%m-%d')
                            end = datetime.strptime(end, '%Y-%m-%d')
                            #get start and end months, and the next month after the start to check if monthly data
                            if int(start.month) <10:
                                month1 = str(start.year)+"/"+'0'+str(start.month)
                            else:
                                month1 = str(start.year)+"/"+str(start.month)
                            if int(start.month)+1 >12:
                                nextmonth=1
                                nextyear=start.year + 1
                            else:
                                nextmonth=start.month+1
                                nextyear=start.year
                            if int(end.month) <10:
                                month2 = str(end.year)+"/"+'0'+str(end.month)
                            else:
                                month2 = str(end.year)+"/"+str(end.month)
                            #checks if not monthly data
                            if nextmonth != end.month:
                                #create start date for the calendarized entry
                                startdate = start
                                startmonth = start.month
                                startday = start.day
                                #checks if you need to vary the year on the date
                                if end.month - start.month < 0:
                                    diff = 12 + (end.month - start.month)
                                else:
                                    diff = end.month - start.month
                                for m in range(0,diff):
                                    if m != 0:
                                        try:
                                            startdate = startdate + timedelta(days=(datetime(nextyear,nextmonth,startday) - startdate).days) #moves to next month
                                        # exception for february
                                        except:
                                            startdate = startdate + timedelta(days=(datetime(nextyear,nextmonth,28) - startdate).days) #moves to next month

                                        startmonth = startmonth + 1
                                        nextmonth = nextmonth + 1
                                    if nextmonth >12:
                                        nextmonth=1
                                        nextyear=start.year + 1
                                    else:
                                        nextyear=start.year
                                    
                                    if startmonth >12:
                                        startmonth=1
                                        startyear=start.year + 1
                                    else:
                                        startyear=start.year
                                        
                                    #calculate total days in each month
                                    daysInMonth1 = (datetime(nextyear,nextmonth,1) - startdate).days
                                    #calculate number of days billed for each month
                                    if daysInMonth1 > startdate.day:
                                        month1Days = daysInMonth1 - startdate.day + 1
                                    else:
                                        month1Days = daysInMonth1
                                    if m == diff:
                                        month2Days = end.day - 1
                                    else: 
                                        month2Days = (startdate + timedelta(days=daysInMonth1)).day - 1
                                    billingPeriod = month1Days+month2Days
                                    
                                    if int(startmonth) <10:
                                        month1 = str(startyear)+"/"+'0'+str(startmonth)
                                    else:
                                        month1 = str(startyear)+"/"+str(startmonth)
                                    if int(nextmonth) <10:
                                        month2 = str(nextyear)+"/"+'0'+str(nextmonth)
                                    else:
                                        month2 = str(nextyear)+"/"+str(nextmonth)
                                    
                                    #get values
                                    if 'cost' in data[acc_id][prop_id]['meters'][mtr_id]['consumptionData'][entry]:
                                        cost = data[acc_id][prop_id]['meters'][mtr_id]['consumptionData'][entry]['cost']
                                    else:
                                        cost = 0
                                    cons = data[acc_id][prop_id]['meters'][mtr_id]['consumptionData'][entry]['value']
                                    #calendarized values
                                    if which_metrics != "Energy" and data[acc_id][prop_id]['meters'][mtr_id]['type'] == 'Municipally Supplied Potable Water - Indoor':
                                        cost1=cost*(month1Days/(billingPeriod*wFreq))
                                        cost2=cost*(month2Days/(billingPeriod*wFreq))
                                        cons1=cons*(month1Days/(billingPeriod*wFreq))
                                        cons2=cons*(month2Days/(billingPeriod*wFreq))                     
                                    else:
                                        cost1=cost*(month1Days/billingPeriod)
                                        cost2=cost*(month2Days/billingPeriod)
                                        cons1=cons*(month1Days/billingPeriod)
                                        cons2=cons*(month2Days/billingPeriod)
                                    #if there is an entry for a given month, add, else create
                                    if month1 in data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"].keys():
                                        data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month1]['Consumption'] += cons1
                                        data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month1]['Cost'] += cost1
                                    else:
                                        data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"].update({month1:{}})    
                                        data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month1].update({'Consumption':cons1,'Cost':cost1})
                                    if month2 in data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"].keys():
                                        data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month2]['Consumption'] += cons2
                                        data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month2]['Cost'] += cost2
                                    else:
                                        data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"].update({month2:{}})    
                                        data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month2].update({'Consumption':cons2,'Cost':cost2})
                        
                            else:
                                #calculate total days in each month
                                daysInMonth1 = (datetime(nextyear,nextmonth,1) - datetime(start.year, start.month,1)).days
                                #calculate number of days billed for each month
                                month1Days = daysInMonth1 - start.day
                                month2Days = end.day
                                billingPeriod = month1Days+month2Days
                                #get values
                                if 'cost' in data[acc_id][prop_id]['meters'][mtr_id]['consumptionData'][entry]:
                                    cost = data[acc_id][prop_id]['meters'][mtr_id]['consumptionData'][entry]['cost']
                                else:
                                    cost = 0
                                cons = data[acc_id][prop_id]['meters'][mtr_id]['consumptionData'][entry]['value']
                                #calendarized values
                                cost1=cost*(month1Days/billingPeriod)
                                cost2=cost*(month2Days/billingPeriod)
                                cons1=cons*(month1Days/billingPeriod)
                                cons2=cons*(month2Days/billingPeriod)
                                #if there is an entry for a given month, add, else create
                                if month1 in data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"].keys():
                                    data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month1]['Consumption'] += cons1
                                    data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month1]['Cost'] += cost1
                                else:
                                    data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"].update({month1:{}})    
                                    data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month1].update({'Consumption':cons1,'Cost':cost1})
                                if month2 in data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"].keys():
                                    data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month2]['Consumption'] += cons2
                                    data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month2]['Cost'] += cost2
                                else:
                                    data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"].update({month2:{}})    
                                    data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][month2].update({'Consumption':cons2,'Cost':cost2})
                            
    return data
#%% DATAFRAMES

def prop_info_to_df(data,acc_id,prop_id,prop1_temp,acc_temp,name_temp,propType_temp,yearBuilt_temp):
    prop1_temp.append(prop_id)
    acc_temp.append(acc_id)
    if data[acc_id][prop_id]:
        name_temp.append(data[acc_id][prop_id]["name"])
        propType_temp.append(data[acc_id][prop_id]["propertyType"])
        yearBuilt_temp.append(data[acc_id][prop_id]["yearBuilt"])
    else:
        name_temp.append("")
        propType_temp.append("")
        yearBuilt_temp.append("")
    

    return prop1_temp,acc_temp,name_temp,propType_temp,yearBuilt_temp

def address_to_df(data,acc_id,prop_id,add1_temp,city_temp,country_temp,postCode_temp,state_temp,prop2_temp):
    prop2_temp.append(prop_id)
    if data[acc_id][prop_id]:
        add1_temp.append(data[acc_id][prop_id]["address"]["address1"])
        city_temp.append(data[acc_id][prop_id]["address"]["city"])
        country_temp.append(data[acc_id][prop_id]["address"]["country"])
        postCode_temp.append(data[acc_id][prop_id]["address"]["postalCode"])
        state_temp.append(data[acc_id][prop_id]["address"]["state"])
    else:
        add1_temp.append("")
        city_temp.append("")
        country_temp.append("")
        postCode_temp.append("")
        state_temp.append("")
    
    return add1_temp,city_temp,country_temp,postCode_temp,state_temp,prop2_temp

def gfa_to_df(data,acc_id,prop_id,gfa_temp,units1_temp,prop3_temp):
    prop3_temp.append(prop_id)
    if data[acc_id][prop_id]:
        gfa_temp.append(data[acc_id][prop_id]["grossFloorArea"]["value"])
        units1_temp.append(data[acc_id][prop_id]["grossFloorArea"]["units"])
    else:
        gfa_temp.append("")
        units1_temp.append("")
        
    return gfa_temp,units1_temp,prop3_temp

def metrics_to_df(data,acc_id,prop_id,year,met_year_temp,score_temp,eui_temp,ghg_temp,prop4_temp):
    prop4_temp.append(prop_id)
    if data[acc_id][prop_id]['metrics'][year]:
        met_year_temp.append(year)
        score_temp.append(data[acc_id][prop_id]["metrics"][year]["score"])
        eui_temp.append(data[acc_id][prop_id]["metrics"][year]["siteIntensity"])
        ghg_temp.append(data[acc_id][prop_id]["metrics"][year]["totalLocationBasedGHGEmissions"])
    else:
        met_year_temp.append("")
        score_temp.append("")
        eui_temp.append("")
        ghg_temp.append("")
    
    return met_year_temp,score_temp,eui_temp,ghg_temp,prop4_temp

def meters_to_df(data,acc_id,prop_id,mtr_id,mtr1_temp,mtrType_temp,units2_temp,prop5_temp):
    prop5_temp.append(prop_id)
    if data[acc_id][prop_id]["meters"]:
        mtr1_temp.append(mtr_id)
        mtrType_temp.append(data[acc_id][prop_id]["meters"][mtr_id]["type"])
        units2_temp.append(data[acc_id][prop_id]["meters"][mtr_id]["units"])
    else:
        mtr1_temp.append("")
        mtrType_temp.append("")
        units2_temp.append("")
    
    return mtr1_temp,mtrType_temp,units2_temp,prop5_temp

def cons_to_df(data,maxEntries,mtr2_temp,month_temp,cons_temp,cost_temp):
    # maybe get rid of new earliest date. instead, count how many entries in each and loop for that many times to add blanks
    for acc_id in data:
        for prop_id in data[acc_id]:
            if data[acc_id][prop_id]:
                for mtr_id in data[acc_id][prop_id]['meters']:
                    count=0
                    if data[acc_id][prop_id]['meters'][mtr_id]:
                        for entry in data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"]:
                            mtr2_temp.append(mtr_id)
                            month_temp.append(entry)
                            cons_temp.append(data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][entry]['Consumption'])
                            if 'Cost' in data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][entry].keys():
                                cost_temp.append(data[acc_id][prop_id]["meters"][mtr_id]["consumptionData_calendarized"][entry]['Cost'])
                            else:
                                cost_temp.append("")
                            count+=1
                        if count!=maxEntries:
                            for x in range(maxEntries-count):
                                mtr2_temp.append("")
                                month_temp.append("")
                                cons_temp.append("")
                                cost_temp.append("")
                    else:
                        for x in range(maxEntries):
                            mtr2_temp.append("")
                            month_temp.append("")
                            cons_temp.append("")
                            cost_temp.append("")
        
    return mtr2_temp,month_temp,cons_temp,cost_temp

#%% CREATE DATAFRAMES
def create_dataframes(prop1_temp,acc_temp,name_temp,propType_temp,yearBuilt_temp,
                      add1_temp,city_temp,country_temp,postCode_temp,state_temp,
                      prop2_temp,gfa_temp,units1_temp,prop3_temp,met_year_temp,
                      score_temp,eui_temp,ghg_temp,prop4_temp,mtr1_temp,mtrType_temp,
                      units2_temp,prop5_temp,mtr2_temp,month_temp,cons_temp,
                      cost_temp):
    dict1 = {'Property ID':prop1_temp, 'Account ID':acc_temp, 'Name':name_temp, 
         'Property Type':propType_temp, 'Year Built':yearBuilt_temp}        
    df1 = pd.DataFrame.from_dict(dict1)
    df1.set_index('Property ID')
    dict2 = {'Street Address':add1_temp, 'City':city_temp, 'Country':country_temp, 
             'Postal Code':postCode_temp, 'State':state_temp, 'Property ID':prop2_temp}        
    df2 = pd.DataFrame.from_dict(dict2)
    df2.set_index('Property ID')
    dict3 = {'GFA':gfa_temp, 'Units':units1_temp, 'Property ID':prop3_temp}        
    df3 = pd.DataFrame.from_dict(dict3)
    dict4 = {'Year':met_year_temp, 'Score':score_temp, 'EUI':eui_temp, 
             'GHG Emissions':ghg_temp, 'Property ID':prop4_temp}        
    df4 = pd.DataFrame.from_dict(dict4)
    dict5 = {'Meter ID':mtr1_temp, 'Meter Type':mtrType_temp, 'Units':units2_temp, 
             'Property ID':prop5_temp}        
    df5 = pd.DataFrame.from_dict(dict5)
    dict6 = {'Meter ID':mtr2_temp, 'Month':month_temp, 'Consumption':cons_temp, 
             'Cost':cost_temp}        
    df6 = pd.DataFrame.from_dict(dict6)
    df6.replace("", float('NaN'), inplace=True)
    df6.dropna(subset=['Meter ID'], inplace=True)
    
    df_info = pd.merge(df1,df2, how='outer', on='Property ID')
    df_info = pd.merge(df_info,df3, how='outer', on='Property ID')
    df_info.set_index('Property ID')
    
    df_metrics = df4
    df_meters = pd.merge(df6,df5, how='outer', on='Meter ID')
    
    return df_info, df_metrics, df_meters

#%% CREATE ENERGY AND WATER DF TO PASTE
def convert_meterdf(prop_id,df_meters,which_metrics):
    
    #converts to kbtu
    electricity_conversion = {'kBtu (thousand Btu)':1,'MBtu/MMBtu (million Btu)':1000,'kWh (thousand Watt-hours)':3.412,'MWh (million Watt-hours)':3412,'GJ':947.817}
    natGas_conversion = {'kBtu (thousand Btu)':1,'MBtu/MMBtu (million Btu)':1000,'cf (cubic feet)':1.026,'ccf (hundred cubic feet)':102.6,'kcf (thousand cubic feet)':1026,'Mcf (million cubic feet)':1026000,'therms':100,'cm (cubic meters)':36.303,'GJ':947.817}
    propane_conversion = {'kBtu (thousand Btu)':1,'MBtu/MMBtu (million Btu)':1000,'cf (cubic feet)':2.516,'ccf (hundred cubic feet)':251.6,'kcf (thousand cubic feet)':2516,'Gallons (US)':92,'Gallons (UK)':110.484,'Liters':24.304,'GJ':947.817}
    districtSteam_conversion = {'kBtu (thousand Btu)':1,'MBtu/MMBtu (million Btu)':1000,'Lbs. (pounds)':1.194,'kLbs. (thousand pounds)':1194,'Mlbs. (million pounds)':1194000,'therms':100,'kg (kilograms)':2.632,'GJ':947.817}
    #converts to gal
    water_conversion = {'ccf (hundred cubic feet)':748.052,'cf (cubic feet)':7.48052,'cGal (hundred gallons) (UK)':120.1,'cGal (hundred gallons) (US)':100,'cm (cubic meters)':264.2,'Gallons (UK)':1.201,'Gallons (US)':1,'kcf (thousand cubic feet)':7481.,'kcm (thousand cubic meters)':264170,'kGal (thousand gallons) (UK)':1201,'kGal (thousand gallons) (US)':1000,'Liters':0.264172, 'Mcf (million cubic feet)':7480520,'MGal (million gallons) (UK)':1201000,'MGal (million gallons) (US)':1000000}
    #preset month matching
    monthNames = {'01':'Jan', '02':'Feb','03':'Mar','04':'Apr','05':'May','06':'Jun','07':'Jul','08':'Aug','09':'Sep','10':'Oct','11':'Nov','12':'Dec'}
    
    #for prop_id, list of meters at that property
    df_meters_temp = df_meters.loc[df_meters['Property ID'] == prop_id]
    
    if which_metrics != "Energy":
        #split to water
        df_meters_temp_water = df_meters_temp.loc[df_meters_temp['Meter Type'].isin(['Municipally Supplied Potable Water - Indoor','Municipally Supplied Potable Water - Outdoor','Municipally Supplied Potable Water - Mixed Indoor/Outdoor'])]
        #year column
        df_meters_temp_water['YEAR'] = df_meters_temp_water['Month'].str[:4]
        #create new cost and meter id column
        df_meters_temp_water['METER ID'] = df_meters_temp_water['Meter ID']
        df_meters_temp_water['COST'] = df_meters_temp_water['Cost']
        df_meters_temp_water['MONTH INDEX'] = df_meters_temp_water['Month']
        #create gal, kgal and new month column
        df_meters_temp_water['POTABLE WATER [GAL]']=0
        df_meters_temp_water['MONTH']=0
        for entry in range(0,len(df_meters_temp_water)):
            df_meters_temp_water['POTABLE WATER [GAL]'].iloc[entry] = df_meters_temp_water['Consumption'].iloc[entry] * water_conversion[df_meters_temp_water['Units'].iloc[entry]]
            df_meters_temp_water['MONTH'].iloc[entry] = monthNames[df_meters_temp_water['Month'].iloc[entry][-2:]]
        #final df to paste    
        df_meters_temp_water=df_meters_temp_water.sort_values(by=['YEAR', 'MONTH INDEX'])
        df_meters_temp_water = df_meters_temp_water[['YEAR', 'MONTH','POTABLE WATER [GAL]','COST','METER ID']]
    else:
        df_meters_temp_water = pd.DataFrame(columns=['YEAR', 'MONTH','POTABLE WATER [GAL]','COST','METER ID'])
        
    #split to energy
    df_meters_temp_energy = df_meters_temp[~df_meters_temp['Meter Type'].isin(['Municipally Supplied Potable Water - Indoor','Municipally Supplied Potable Water - Outdoor','Municipally Supplied Potable Water - Mixed Indoor/Outdoor'])]
    #year column
    df_meters_temp_energy['YEAR'] = df_meters_temp_energy['Month'].str[:4]
    #create new cost, meter type and meter id column
    df_meters_temp_energy['METER ID'] = df_meters_temp_energy['Meter ID']
    df_meters_temp_energy['COST'] = df_meters_temp_energy['Cost']
    df_meters_temp_energy['ENERGY SOURCE'] = df_meters_temp_energy['Meter Type'] 
    df_meters_temp_energy['MONTH INDEX'] = df_meters_temp_energy['Month']
    
    #Create kbtu and new month column
    df_meters_temp_energy['KBTU']=0
    df_meters_temp_energy['MONTH']=0
    for entry in range(0,len(df_meters_temp_energy)):
        df_meters_temp_energy['MONTH'].iloc[entry] = monthNames[df_meters_temp_energy['Month'].iloc[entry][-2:]]
        if df_meters_temp_energy['Meter Type'].iloc[entry] == "Electric":
            df_meters_temp_energy['KBTU'].iloc[entry] = df_meters_temp_energy['Consumption'].iloc[entry] * electricity_conversion[df_meters_temp_energy['Units'].iloc[entry]]
        elif df_meters_temp_energy['Meter Type'].iloc[entry] == "Electric on Site Solar":
            df_meters_temp_energy['KBTU'].iloc[entry] = df_meters_temp_energy['Consumption'].iloc[entry] * electricity_conversion[df_meters_temp_energy['Units'].iloc[entry]]
        elif df_meters_temp_energy['Meter Type'].iloc[entry] == "Natural Gas":
            df_meters_temp_energy['KBTU'].iloc[entry] = df_meters_temp_energy['Consumption'].iloc[entry] * natGas_conversion[df_meters_temp_energy['Units'].iloc[entry]]
        elif df_meters_temp_energy['Meter Type'].iloc[entry] == "Propane":
            df_meters_temp_energy['KBTU'].iloc[entry] = df_meters_temp_energy['Consumption'].iloc[entry] * propane_conversion[df_meters_temp_energy['Units'].iloc[entry]]
        elif df_meters_temp_energy['Meter Type'].iloc[entry] == "District Steam":
            df_meters_temp_energy['KBTU'].iloc[entry] = df_meters_temp_energy['Consumption'].iloc[entry] * electricity_conversion[districtSteam_conversion['Units'].iloc[entry]]    
    #final df to paste    
    df_meters_temp_energy=df_meters_temp_energy.sort_values(by=['YEAR', 'MONTH INDEX'])
    df_meters_temp_energy = df_meters_temp_energy[['YEAR', 'MONTH','KBTU','COST','ENERGY SOURCE','METER ID']]
    return df_meters_temp_energy, df_meters_temp_water

#%%BUILD TARGETS TABLES
    
def build_targetsdf(prop_id,inputs_df,df_info_temp,df_meters_temp_energy,df_meters_temp_water,which_metrics,which_properties,transpTotal,baselineE,baselineW,baselineT):
    #energy
    yearsE = sorted(df_meters_temp_energy.YEAR.unique())
    df_noRE = df_meters_temp_energy[df_meters_temp_energy['ENERGY SOURCE'] != 'Electric on Site Solar']
    df_energySum = df_noRE.groupby('YEAR').sum()
    if which_properties == 'District Aggregate':
        countE = df_noRE.groupby(["YEAR","ENERGY SOURCE"]).size()
    else:
        countE = df_noRE.groupby(["YEAR","ENERGY SOURCE","METER ID"]).size()
    df_RE = df_meters_temp_energy[df_meters_temp_energy['ENERGY SOURCE'] == 'Electric on Site Solar']
    df_RESum = df_RE.groupby(['YEAR','ENERGY SOURCE']).sum()
    countRE = df_RE.groupby(['YEAR','ENERGY SOURCE']).size()
    #calc values
    if which_properties == 'District Aggregate':
        baselineE = baselineE
    else:
        try:
            baselineE = int(inputs_df.loc[inputs_df.index[inputs_df['Property ID (from ESPM Profile)'] == prop_id][0],'Baseline EUI [kBtu/ft2]'])
        except:
            ctypes.windll.user32.MessageBoxW(0, "Check the inputs sheet - your baseline EUI value was not accepted.","Error", 0)
    Tar2020E = baselineE*.8
    Tar2025E = baselineE*.65
    Tar2030E = baselineE*.5
    eui = {}
    # ((12)/(#entries))x((#sumKBTU-RE_KBTU/SF)) = EUI
    for year in yearsE:
        #if year not in countRE:
        #    eui.update({year:(((12/max(countE[year]))*float(df_energySum.KBTU[year])))/float(df_info_temp.iloc[0]['GFA'])})
        #elif year not in countE:
        #    eui.update({year:(-((12/max(countRE[year]))*float(df_RESum.KBTU[year])))/float(df_info_temp.iloc[0]['GFA'])})
        #else:
        #    eui.update({year:(((12/max(countE[year]))*float(df_energySum.KBTU[year])) - ((12/max(countRE[year]))*float(df_RESum.KBTU[year])))/float(df_info_temp.iloc[0]['GFA'])})
        if year in countE:
            eui.update({year:(((12/max(countE[year]))*float(df_energySum.KBTU[year])))/float(df_info_temp.iloc[0]['GFA'])})
        else:
            eui.update({year:(0)})
        
    dictTempE={}
    dictTempE={'Vertical':[],'Horizontal':[],'Metric':[]}

    horiz = []
    horiz = list(yearsE)
    horiz.append('Baseline')
    dictTempE['Horizontal'] = horiz * 4
    vert = []
    metric = []

    for year in yearsE:
        metric.append(eui[year])
        if year == max(yearsE):
            vert.append('Current')
        else:
            vert.append('Historic')
    vert.append('Baseline')
    metric.append(baselineE)
    for count in range(0,len(yearsE)+1):
        vert.append('2020 Reduction Target')
        metric.append(Tar2020E)
    for count in range(0,len(yearsE)+1):
        vert.append('2025 Reduction Target')
        metric.append(Tar2025E)
    for count in range(0,len(yearsE)+1):
        vert.append('2030 Reduction Target')
        metric.append(Tar2030E)
    
    dictTempE['Vertical'] = vert
    dictTempE['Metric'] = metric
    df_targets_energy = pd.DataFrame(dictTempE)

    #water
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        yearsW = sorted(df_meters_temp_water.YEAR.unique())
        if which_properties == 'District Aggregate':
            countW = df_meters_temp_water.groupby(["YEAR"]).size()
        else:
            countW = df_meters_temp_water.groupby(["YEAR","METER ID"]).size()
        #calc values
        if which_properties == 'District Aggregate':
            baselineW = baselineW
        else:
            try:
                baselineW = int(inputs_df.loc[inputs_df.index[inputs_df['Property ID (from ESPM Profile)'] == prop_id][0],'Baseline WUI [gal/ft2]'])
            except:
                ctypes.windll.user32.MessageBoxW(0, "Check the inputs sheet - your baseline WUI value was not accepted.","Error", 0)
        Tar2020W = baselineW*.8
        Tar2025W = baselineW*.65
        Tar2030W = baselineW*.5
        wui = {}
        
        for year in yearsW:
            try:
                wui.update({year:(12/max(countW[year]))*float(df_meters_temp_water.loc[df_meters_temp_water['YEAR'] == year]["POTABLE WATER [GAL]"].sum())/float(df_info_temp.iloc[0]['GFA'])})
            except:  
                wui.update({year:(12/(countW[year]))*float(df_meters_temp_water.loc[df_meters_temp_water['YEAR'] == year]["POTABLE WATER [GAL]"].sum())/float(df_info_temp.iloc[0]['GFA'])})                 
                                                                        
        dictTempW={}
        dictTempW={'Vertical':[],'Horizontal':[],'Metric':[]}
    
        horiz = []
        horiz = list(yearsW)
        horiz.append('Baseline')
        dictTempW['Horizontal'] = horiz * 4
        vert = []
        metric = []
        
        for year in yearsW:
            metric.append(wui[year])
            if year == max(yearsW):
                vert.append('Current')
            else:
                vert.append('Historic')
        vert.append('Baseline')
        metric.append(baselineW) 
        for count in range(0,len(yearsW)+1):
            vert.append('2020 Reduction Target')
            metric.append(Tar2020W)
        for count in range(0,len(yearsW)+1):
            vert.append('2025 Reduction Target')
            metric.append(Tar2025W)
        for count in range(0,len(yearsW)+1):
            vert.append('2030 Reduction Target')
            metric.append(Tar2030W)
    
        dictTempW['Vertical'] = vert
        dictTempW['Metric'] = metric
        
        df_targets_water = pd.DataFrame(dictTempW)
    else:
        df_targets_water = pd.DataFrame(columns=['Vertical','Horizontal','Metric'])
        
    df_targets_energy = pd.DataFrame(dictTempE)
    
    #transportation
    if which_metrics == "Energy & Water & Transportation" and transpTotal.empty and "Ithaca" in district_name:
        ctypes.windll.user32.MessageBoxW(0, "Check the inputs sheet - no data was input for Total Annual Emissions.","Error", 0)
    elif which_metrics == "Energy & Water & Transportation" and not transpTotal.empty and "Ithaca" in district_name:
        yearsT = sorted(transpTotal['Year.1'])
        df_TSum = transpTotal.groupby('Year.1').sum()

        #calc values
        Tar2020T = baselineT*.8
        Tar2025T = baselineT*.65
        Tar2030T = baselineT*.5
        emissions = {}
        # ((12)/(#entries))x((#sumKBTU-RE_KBTU/SF)) = EUI
        for year in yearsT:
            emissions.update({year:df_TSum.loc[year][0]})
            
        dictTempT={}
        dictTempY={'Vertical':[],'Horizontal':[],'Metric':[]}
    
        horiz = []
        horiz = list(yearsT)
        horiz.append('Baseline')
        dictTempT['Horizontal'] = horiz * 4
        vert = []
        metric = []
    
        for year in yearsT:
            metric.append(emissions[year])
            if year == max(yearsT):
                vert.append('Current')
            else:
                vert.append('Historic')
        vert.append('Baseline')
        metric.append(baselineT)
        for count in range(0,len(yearsT)+1):
            vert.append('2020 Reduction Target')
            metric.append(Tar2020T)
        for count in range(0,len(yearsT)+1):
            vert.append('2025 Reduction Target')
            metric.append(Tar2025T)
        for count in range(0,len(yearsT)+1):
            vert.append('2030 Reduction Target')
            metric.append(Tar2030T)
        
        dictTempT['Vertical'] = vert
        dictTempT['Metric'] = metric
        
        df_targets_transp = pd.DataFrame(dictTempT)
        df_targets_transp = df_targets_transp.reindex(['Vertical','Horizontal','Metric'],axis=1)
    
    if which_metrics != "Energy & Water & Transportation" or "Ithaca" not in district_name:
        df_targets_transp = pd.DataFrame()
        
    return df_targets_energy,df_targets_water,countRE,df_targets_transp



#%% COPY PASTE BETWEEN SHEETS
    
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
        




#%% CREATE AND FORMAT TABLES


def create_tables(df_meters_temp_energy,df_meters_temp_water,df_info_temp,
                  df_metrics_temp,df_targets_energy,df_targets_water,df_targets_transp,which_metrics,which_properties,prop_id):
    """
    """
    #paste data to temp workbooks
    df_meters_temp_energy.to_excel('energyTemp.xlsx',index=False) 
    if which_metrics != "Energy" and not df_meters_temp_water.empty:    
        df_meters_temp_water.to_excel('waterTemp.xlsx',index=False) 
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        transpSplit.to_excel('transpTemp.xlsx',index=False)
    df_info_temp.to_excel('infoTemp.xlsx',index=False)
    df_metrics_temp.to_excel('metricsTemp.xlsx',index=False)
    df_targets_energy.to_excel('targetsTempE.xlsx',index=True)
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        df_targets_water.to_excel('targetsTempW.xlsx',index=True)
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        df_targets_transp.to_excel('targetsTempT.xlsx',index=True)
    
    #open temp workbooks
    wb_temp_energy = load_workbook('energyTemp.xlsx')
    ws_temp_energy = wb_temp_energy.active
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        wb_temp_water = load_workbook('waterTemp.xlsx')
        ws_temp_water = wb_temp_water.active
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        wb_temp_transp = load_workbook('transpTemp.xlsx')
        ws_temp_transp = wb_temp_transp.active
    wb_temp_info = load_workbook('infoTemp.xlsx')
    ws_temp_info = wb_temp_info.active
    wb_temp_metrics = load_workbook('metricsTemp.xlsx')
    ws_temp_metrics = wb_temp_metrics.active
    wb_temp_targetsE = load_workbook('targetsTempE.xlsx')
    ws_temp_targetsE = wb_temp_targetsE.active
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        wb_temp_targetsW = load_workbook('targetsTempW.xlsx')
        ws_temp_targetsW = wb_temp_targetsW.active
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        wb_temp_targetsT = load_workbook('targetsTempT.xlsx')
        ws_temp_targetsT = wb_temp_targetsT.active
    
    
    #create new workbook that has all data
    wb = Workbook()
    wsE = wb.create_sheet('Energy')
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        wsW = wb.create_sheet('Water')
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        wsT = wb.create_sheet('Transp')
    wsI = wb.create_sheet('Info')
    wsM = wb.create_sheet('Metrics')
    wsTE = wb.create_sheet('TargetsE')
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        wsTW = wb.create_sheet('TargetsW')
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        wsTT = wb.create_sheet('TargetsT')
    wb.remove(wb['Sheet'])
    
    #copy and paste from separate workbooks into one workbook
    selectedRangeE = copyRange(1,1,df_meters_temp_energy.shape[1],df_meters_temp_energy.shape[0]+1,ws_temp_energy)
    pasteRange(1,1,df_meters_temp_energy.shape[1],df_meters_temp_energy.shape[0]+1,wsE,selectedRangeE)
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        selectedRangeW = copyRange(1,1,df_meters_temp_water.shape[1],df_meters_temp_water.shape[0]+1,ws_temp_water)
        pasteRange(1,1,df_meters_temp_water.shape[1],df_meters_temp_water.shape[0]+1,wsW,selectedRangeW)
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        selectedRangeT = copyRange(1,1,transpSplit.shape[1],transpSplit.shape[0]+1,ws_temp_transp)
        pasteRange(1,1,transpSplit.shape[1],transpSplit.shape[0]+1,wsT,selectedRangeT)
    selectedRangeI = copyRange(1,1,df_info_temp.shape[1],df_info_temp.shape[0]+1,ws_temp_info)
    pasteRange(1,1,df_info_temp.shape[1],df_info_temp.shape[0]+1,wsI,selectedRangeI)
    selectedRangeM = copyRange(1,1,df_metrics_temp.shape[1],df_metrics_temp.shape[0]+1,ws_temp_metrics)
    pasteRange(1,1,df_metrics_temp.shape[1],df_metrics_temp.shape[0]+1,wsM,selectedRangeM)
    selectedRangeTE = copyRange(2,1,df_targets_energy.shape[1]+1,df_targets_energy.shape[0]+1,ws_temp_targetsE)
    pasteRange(1,1,df_targets_energy.shape[1],df_targets_energy.shape[0]+1,wsTE,selectedRangeTE)
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        selectedRangeTW = copyRange(2,1,df_targets_water.shape[1]+1,df_targets_water.shape[0]+1,ws_temp_targetsW)
        pasteRange(1,1,df_targets_water.shape[1],df_targets_water.shape[0]+1,wsTW,selectedRangeTW)
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        selectedRangeTT = copyRange(2,1,df_targets_transp.shape[1]+1,df_targets_transp.shape[0]+1,ws_temp_targetsT)
        pasteRange(1,1,df_targets_transp.shape[1],df_targets_transp.shape[0]+1,wsTT,selectedRangeTT)
    #SAVE
    wb.save(prop_id+".xlsx")
    
    #create table styles
    if which_properties == "District Aggregate":
        ref_temp_energy='A1:E'+str(1+df_meters_temp_energy.shape[0])
        ref_temp_metrics='A1:D'+str(1+df_metrics_temp.shape[0])
        if which_metrics != "Energy" and not df_meters_temp_water.empty:
            ref_temp_water='A1:D'+str(1+df_meters_temp_water.shape[0])
            tab_temp_water = Table(displayName='WATERDATA',ref=ref_temp_water)
        if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
            ref_temp_transp='A1:E'+str(1+transpSplit.shape[0])
            tab_temp_transp = Table(displayName='TRANSPDATA',ref=ref_temp_transp)
            
    else:
        ref_temp_energy='A1:F'+str(1+df_meters_temp_energy.shape[0])
        ref_temp_metrics='A1:E'+str(1+df_metrics_temp.shape[0])
        if which_metrics != "Energy" and not df_meters_temp_water.empty:
            ref_temp_water='A1:E'+str(1+df_meters_temp_water.shape[0])
            tab_temp_water = Table(displayName='WATERDATA',ref=ref_temp_water)
        if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
            ref_temp_transp='A1:E'+str(1+transpSplit.shape[0])
            tab_temp_transp = Table(displayName='TRANSPDATA',ref=ref_temp_transp)
    
    tab_temp_energy = Table(displayName='ENERGYDATA',ref=ref_temp_energy)
    tab_temp_metrics = Table(displayName='METRICS',ref=ref_temp_metrics)
    
    ref_temp_TE ='A1:'+str(chr(ord('A')+(df_targets_energy.shape[1])-1))+str(1+df_targets_energy.shape[0])
    tab_temp_TE = Table(displayName='ENERGYTARGETDATA',ref=ref_temp_TE)
    
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        ref_temp_TW='A1:'+str(chr(ord('A')+(df_targets_water.shape[1])-1))+str(1+df_targets_water.shape[0])
        tab_temp_TW = Table(displayName='WATERTARGETDATA',ref=ref_temp_TW)
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        ref_temp_TT='A1:'+str(chr(ord('A')+(df_targets_transp.shape[1])-1))+str(1+df_targets_transp.shape[0])
        tab_temp_TT = Table(displayName='TRANSPTARGETDATA',ref=ref_temp_TT)
    
    style = TableStyleInfo(name='TableStyleMedium2',showFirstColumn=False,
                           showLastColumn=False,showRowStripes=True,showColumnStripes=False)
    #save temp tables
    tab_temp_energy.tableStyleInfo = style
    wsE.add_table(tab_temp_energy)
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        tab_temp_water.tableStyleInfo = style
        wsW.add_table(tab_temp_water)
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        tab_temp_transp.tableStyleInfo = style
        wsT.add_table(tab_temp_transp)
    tab_temp_metrics.tableStyleInfo = style
    wsM.add_table(tab_temp_metrics)
    tab_temp_TE.tableStyleInfo = style
    wsTE.add_table(tab_temp_TE)
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        tab_temp_TW.tableStyleInfo = style
        wsTW.add_table(tab_temp_TW)
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        tab_temp_TT.tableStyleInfo = style
        wsTT.add_table(tab_temp_TT)
    wb.save(prop_id+'.xlsx')
    
    #remove temp files
    os.remove("energyTemp.xlsx")
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        os.remove("waterTemp.xlsx")
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        os.remove("transpTemp.xlsx")
    os.remove("infoTemp.xlsx")
    os.remove("metricsTemp.xlsx")
    os.remove("targetsTempE.xlsx")
    if which_metrics != "Energy" and not df_meters_temp_water.empty:
        os.remove("targetsTempW.xlsx")
    if which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name:
        os.remove("targetsTempT.xlsx")
    
    #kill excel
    os.system("taskkill /f /im  Excel.exe")
    

#%% RUN SCRIPT

os.system("taskkill /f /im  Excel.exe")

# GET PRECONDITIONS
[year_curr,month_curr,day_curr] = get_current_date()
[metrics_list,api_url_base,username,pswd,prop1_temp,acc_temp,name_temp,
 propType_temp,yearBuilt_temp,add1_temp,city_temp,country_temp,postCode_temp,
 state_temp,prop2_temp,gfa_temp,units1_temp,prop3_temp,met_year_temp,score_temp,
 eui_temp,ghg_temp,prop4_temp,mtr1_temp,mtrType_temp,units2_temp,prop5_temp,
 mtr2_temp,month_temp,cons_temp,cost_temp,maxEntries] = set_preconditions()
#preset transportation values to avoid errors
baselineT = pd.DataFrame()
transpSplit = pd.DataFrame()
transpTotal = pd.DataFrame()

#% API Calls

# Get list of pending account, property, and meter requests and accept them
acc_rqsts = xmltodict_acc_rqsts(api_url_base,username,pswd)
accept_acct_requests(api_url_base,acc_rqsts,username,pswd)

prop_rqsts = xmltodict_prop_rqsts(api_url_base,username,pswd)
accept_prop_requests(api_url_base,prop_rqsts,username,pswd)

mtr_rqsts = xmltodict_mtr_rqsts(api_url_base,username,pswd)
accept_mtr_requests(api_url_base,mtr_rqsts,username,pswd)


# Get connected account info and inputs
accounts = xmltodict_acc_cnxn(api_url_base,username,pswd)
inputs_df = get_inputs()


accounts = xmltodict_acc_info(api_url_base,username,pswd,accounts)

district_name = get_district_name()
which_properties = get_which_properties()
which_metrics = get_which_metrics()
[district_eui, district_wui] = get_district_eui_wui()

if which_properties == "District Aggregate":
    # create empty dataframes, iteratively fill and then aggregate at end
    metrics_agg = pd.DataFrame(columns=['Year','Score','EUI','GHG Emissions','Property ID','GFA'])
    consE_agg = pd.DataFrame(columns=['YEAR','MONTH','KBTU','COST','ENERGY SOURCE','METER ID'])
    consW_agg = pd.DataFrame(columns=['YEAR','MONTH','POTABLE WATER [GAL]','COST','METER ID'])

#Stop from running transportation if not Ithaca, otherwise get inputs for agg transportation
if which_metrics == "Energy & Water & Transportation" and "Ithaca" not in district_name:
    ctypes.windll.user32.MessageBoxW(0, "Transportation has not yet been set up for your district. Please speak to developer about template creation.","Error", 0)
    exit()
elif which_metrics == "Energy & Water & Transportation" and "Ithaca" in district_name and which_properties == "District Aggregate":
    [baselineT,transpSplit,transpTotal,new_which_metrics] = get_transp_agg_inputs()

#set up data dictionary with the needed account ids
data = xmltodict_admin_acc(api_url_base,username,pswd,inputs_df)
    
#stop from running if not connected to their account
if not data: 
    ctypes.windll.user32.MessageBoxW(0, "Your account was not found. Check the spelling of your username, and if you are connected to namnum_dashboard on ES Portfolio Manager.","Error", 0)
    exit()
else:
    #% CREATE FOLDER FOR DATA

    #get today's date for folder name
    today = str(datetime.now().month) + "-" + str(datetime.now().day) + "-" + str(datetime.now().year)
    #get current working directory
    cwd = os.getcwd()
    #create new data directory
    #dir = os.path.join(cwd,"Data " + today)
    dir = os.path.join(cwd, "Data")
    if not os.path.exists(dir):
        os.mkdir(dir)
    #go to new directory
    os.chdir(dir)

    for acc_id in data:

        if which_properties != "Individual" and which_properties != "District Aggregate": #get only selected properties
            ctypes.windll.user32.MessageBoxW(0, "Check your answer on the inputs sheet to 'Create individual dashboards for each property, one District Aggregate?'. Your response should either be 'Individual' or 'District Aggregate'.","Error", 0)
            exit()
        
        for prop_id in data[acc_id]:    
            # From connected properties get property info
            try:
                data = xmltodict_prop_info(api_url_base,username,pswd,data,prop_id)
            except:
                ctypes.windll.user32.MessageBoxW(0, "An inputted property ID was not found on ESPM. Make sure it is connected to the namnum_dashboard account.","Error", 0)
                exit()
            # RELATIONAL DB 1,2,3
            prop_info_to_df(data,acc_id,prop_id,prop1_temp,acc_temp,name_temp,
                            propType_temp,yearBuilt_temp)
            address_to_df(data,acc_id,prop_id,add1_temp,city_temp,country_temp,
                          postCode_temp,state_temp,prop2_temp)
            gfa_to_df(data,acc_id,prop_id,gfa_temp,units1_temp,prop3_temp)
            
            # Reset earliestDate for each property
            earliestDate = year_curr + '-' + month_curr + '-' + day_curr
            year_end = reset_date(earliestDate)
            
            if data[acc_id][prop_id]:
                # If a property is not empty, get connected meters
                data = xmltodict_mtr_cnxn(api_url_base,username,pswd,data,
                                               acc_id,prop_id)
                if data[acc_id][prop_id]['meters']:
                    # If meters exist, get meter info and consumption
                    for mtr_id in data[acc_id][prop_id]['meters']:
                        data = xmltodict_mtr_info(api_url_base,username,pswd,
                                                       data,acc_id,prop_id,mtr_id)
                        # Create temp lists to fill meter df
                        meters_to_df(data,acc_id,prop_id,mtr_id,mtr1_temp,
                                     mtrType_temp,units2_temp,prop5_temp)
                        # Replace earliestDate if it is later than the first bill date
                        if str(data[acc_id][prop_id]['meters'][mtr_id]
                               ['firstBillDate'])<earliestDate:
                            earliestDate = str(data[acc_id][prop_id]['meters']
                                               [mtr_id]['firstBillDate'])
                        # Get consumption data using API
                        data = xmltodict_consumption(api_url_base,username,pswd,data,acc_id,prop_id,mtr_id)
                        if len(data[acc_id][prop_id]['meters'][mtr_id]['consumptionData'])>maxEntries:
                            maxEntries = len(data[acc_id][prop_id]['meters'][mtr_id]['consumptionData'])
                #
                year_start = reset_date(earliestDate)
                years = []
                for n in range(int(year_start),int(year_end)+1,1):
                    years.append(str(n)) 
                for year in years:
                    data[acc_id][prop_id]["metrics"].update({year:{}})
                    for metric in metrics_list:
                        # From connected properties get metrics
                        data = xmltodict_metrics(api_url_base,username,pswd,
                                                      data,acc_id,prop_id,year,metric)
                    
                    # RELATIONAL DB 4
                    metrics_to_df(data,acc_id,prop_id,year,met_year_temp,score_temp,
                                  eui_temp,ghg_temp,prop4_temp)
    
    
    data = calendarize_consumption(data,inputs_df,which_metrics)
    # create temp lists to construct df
    cons_to_df(data,maxEntries,mtr2_temp,month_temp,cons_temp,cost_temp)
    
    #% CREATE DATAFRAMES
    [df_info,df_metrics,df_meters] = create_dataframes(prop1_temp,acc_temp,
                                                            name_temp,propType_temp,
                                                            yearBuilt_temp,add1_temp,
                                                            city_temp,country_temp,
                                                            postCode_temp,state_temp,
                                                            prop2_temp,gfa_temp,
                                                            units1_temp,prop3_temp,
                                                            met_year_temp,score_temp,
                                                            eui_temp,ghg_temp,prop4_temp,
                                                            mtr1_temp,mtrType_temp,
                                                            units2_temp,prop5_temp,
                                                            mtr2_temp,month_temp,
                                                            cons_temp,cost_temp)                 
    
    #% CREATING NEW DATA WORKBOOKS
    
    #property counts
    
    propCountE = 0
    propCountW = 0
    
    if which_properties == "District Aggregate":
        for acc_id in data:    
            for prop_id in data[acc_id]:
                #add to energy count for each property
                propCountE = propCountE + 1
                df_info_temp = df_info.loc[df_info['Property ID'] == prop_id]
                
                df_metrics_temp = df_metrics.loc[df_metrics['Property ID'] == prop_id]
                metrics_agg = metrics_agg.append(df_metrics_temp)
                metrics_agg = metrics_agg.fillna(int(df_info_temp.iloc[0]['GFA']))
                #split and convert consumption data
                [df_meters_temp_energy,df_meters_temp_water] = convert_meterdf(prop_id,df_meters,which_metrics)
                consE_agg = consE_agg.append(df_meters_temp_energy)
                consW_agg = consW_agg.append(df_meters_temp_water)
                #drop first & last row of water df, since it won't be accurate unless it's monthly
                if which_metrics != "Energy" and not df_meters_temp_water.empty:
                    #add to water count for each property with water data
                    propCountW = propCountW + 1
                    if inputs_df.loc[inputs_df.index[inputs_df['Property ID (from ESPM Profile)'] == prop_id][0],'Frequency of Water Bills'] != 'Monthly':
                        df_meters_temp_water.drop(df_meters_temp_water.tail(1).index,inplace=True)
            
              
            metrics_agg = metrics_agg.replace("N/A",0)
            metrics_agg['GHG Emissions'] =  metrics_agg['GHG Emissions'].apply(pd.to_numeric)
            metrics_agg['EUI'] =  metrics_agg['EUI'].apply(pd.to_numeric)
            count = metrics_agg.groupby('Year').size().to_frame('count').reset_index()
            metrics_agg = metrics_agg.groupby('Year',as_index=False).sum()
            #calc avg EUI
            metrics_agg['EUI'] = metrics_agg['EUI']/(count['count'].apply(pd.to_numeric))
            metrics_agg = metrics_agg.replace(0,"N/A")
            
            info_agg = pd.DataFrame(columns=['Name','GFA','Units','# Buildings (E)','# Buildings (W)']).append({'Name':district_name,"GFA":max(metrics_agg['GFA']),"Units":df_info_temp.iloc[0]['Units'],"# Buildings (E)":propCountE,"# Buildings (W)":propCountW},ignore_index=True)
            
            consE_agg = consE_agg.groupby(['YEAR','MONTH','ENERGY SOURCE'],as_index=False).sum()
            consW_agg = consW_agg.groupby(['YEAR','MONTH'],as_index=False).sum()
            
            [df_targets_energy,df_targets_water,countRE,df_targets_transp] = build_targetsdf(prop_id,inputs_df,df_info_temp,df_meters_temp_energy,df_meters_temp_water,which_metrics,which_properties,transpTotal,district_eui,district_wui,baselineT)
       
             #change years to numbers
            consE_agg['YEAR'] =  consE_agg['YEAR'].apply(pd.to_numeric)
            if which_metrics != "Energy" and not df_meters_temp_water.empty:
                consW_agg['YEAR'] =  consW_agg['YEAR'].apply(pd.to_numeric)   
            
        #individual dashboard data
    else:
        for acc_id in data:
            for prop_id in data[acc_id]:
                #reset which_metrics
                new_which_metrics = which_metrics
                if data[acc_id][prop_id]:


                    df_info_temp = df_info.loc[df_info['Property ID'] == prop_id]
                    df_metrics_temp = df_metrics.loc[df_metrics['Property ID'] == prop_id]
                    [df_meters_temp_energy,df_meters_temp_water] = convert_meterdf(prop_id,df_meters,new_which_metrics)
                    #change years to numbers
                    df_meters_temp_energy['YEAR'] =  df_meters_temp_energy['YEAR'].apply(pd.to_numeric)
                    df_meters_temp_water['YEAR'] =  df_meters_temp_water['YEAR'].apply(pd.to_numeric)
                    #drop last row of water df, since it won't be accurate unless it's monthly
                    if new_which_metrics != "Energy" and not df_meters_temp_water.empty:
                        if inputs_df.loc[inputs_df.index[inputs_df['Property ID (from ESPM Profile)'] == prop_id][0],'Frequency of Water Bills'] != 'Monthly':
                            df_meters_temp_water.drop(df_meters_temp_water.tail(1).index,inplace=True)

                    [df_targets_energy,df_targets_water,countRE,df_targets_transp] = build_targetsdf(prop_id,inputs_df,df_info_temp,df_meters_temp_energy,df_meters_temp_water,new_which_metrics,which_properties,transpTotal,district_eui,district_wui,baselineT)
                    
                    
                    #% EXCEL FORMATTING
                    create_tables(df_meters_temp_energy,df_meters_temp_water,df_info_temp,
                                  df_metrics_temp,df_targets_energy,df_targets_water,
                                  df_targets_transp,new_which_metrics,which_properties,prop_id)
    
                os.system("taskkill /f /im  Excel.exe")
    
    
    #district aggregate dashboard data
    if which_properties == "District Aggregate":
        prop_id = "Aggregate"
        
        #% EXCEL FORMATTING
        create_tables(consE_agg,consW_agg,info_agg,metrics_agg,df_targets_energy,
                      df_targets_water,df_targets_transp,which_metrics,which_properties,prop_id)
    
        os.system("taskkill /f /im  Excel.exe")
        
    ctypes.windll.user32.MessageBoxW(0, "Data download success.","Congratulations!", 0)
# %%
